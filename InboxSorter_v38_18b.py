# -*- coding: utf-8 -*-
"""Email Sorter Python Application — Version 38.18b

Changes in v38.18:
  - Removed unused imports (simpledialog, duplicate load_workbook)
  - Fixed invalid_logger level to WARNING so .warning() calls are captured
  - Added threading.Lock for smtp_cache / new_smtp_entries (thread safety)
  - save_smtp_cache now snapshots new entries under lock before file I/O
    and clears only successfully saved entries afterwards
  - smtp_fallback_dir stored as self.smtp_fallback_dir in setup_paths
  - _load_email_addresses now strips leading/trailing whitespace
  - Replaced manual HTML entity replacements with html.unescape()
  - extract_sender_address_only uses set literal {x} instead of set([x])
  - Outlook folder magic numbers replaced with named class constants
  - load_data refactored using _load_rule() helper (eliminates 11 repeated blocks)
  - MSIP label constants indentation and comments corrected
  - process_email inspector-check comment indentation fixed
  - General comment clarification throughout

Changes in v38.18b:
  - _inject_msip_label: added _rels/.rels relationship fix — openpyxl silently
    drops the classification relationship on every save, causing Office to ignore
    the label even though LabelInfo.xml is present in the zip; restored via new
    _ensure_label_relationship() helper
  - Added RELS_PATH, LABEL_REL_TYPE, LABEL_REL_TARGET class constants
  - save_smtp_cache: resizes the SMTPResolutionCache Excel Table after appending
    rows so the new entries fall inside the table (auto-filter, formatting, etc.)
  - save_smtp_cache: fixed in-place _inject_msip_label(tmp, tmp) call — src and
    dst cannot safely be the same path; now uses a second tmp file (tmp2) with
    an atomic move, matching the safe pattern from v38.17d
"""

import os
import win32com.client
import pandas as pd
import datetime
import openpyxl
import tkinter as tk
from tkinter import messagebox
from tkcalendar import Calendar
import threading
import logging
import time
import json
import re
import shutil
import pythoncom
import zipfile
from io import BytesIO

class EmailSorter:
    """
    A class to sort emails in Outlook based on rules defined in an Excel file
    and configuration from a JSON file.
    Supports live monitoring and bulk processing modes.
    Version 38.18b: See module docstring for full change list.
    """

    # Define config file path once here for consistency
    CONFIG_FILE_NAME = 'config_v38.10.json'

    # MSIP sensitivity label constants — extracted from docMetadata/LabelInfo.xml
    # inside the smtp cache xlsx. Hardcoded as all fields are fixed for this tenant.
    LABEL_INFO_XML = (
        '<?xml version="1.0" encoding="utf-8" standalone="yes"?>'
        '<clbl:labelList xmlns:clbl="http://schemas.microsoft.com/office/2020/mipLabelMetadata">'
        '<clbl:label'
        ' id="{840e60c6-cef6-4cc0-a98d-364c7249d74b}"'
        ' enabled="1"'
        ' method="Privileged"'
        ' siteId="{b44900f1-2def-4c3b-9ec6-9020d604e19e}"'
        ' removed="0"'
        ' />'
        '</clbl:labelList>'
    )

    # Zip part path and content-type as registered in [Content_Types].xml
    LABEL_PART_PATH    = "docMetadata/LabelInfo.xml"
    LABEL_CONTENT_TYPE = "application/vnd.ms-office.classificationlabels+xml"
    CONTENT_TYPES_PATH = "[Content_Types].xml"
    # Root relationships file — openpyxl drops the classification entry here on every save
    RELS_PATH       = "_rels/.rels"
    LABEL_REL_TYPE  = "http://schemas.microsoft.com/office/2020/02/relationships/classificationlabels"
    LABEL_REL_TARGET = "docMetadata/LabelInfo.xml"

    # Outlook default folder IDs used with GetDefaultFolder()
    _OL_FOLDER_INBOX = 6
    _OL_FOLDER_SENT  = 5

    def __init__(self, config_path=None):
        """
        Initializes the EmailSorter with configuration, sets up paths,
        loads data from Excel.
        """
        # Use provided config_path or the default class-level constant
        self.config_path = config_path if config_path else self.CONFIG_FILE_NAME

        self.config = None
        self.xls_path = None
        self.smtp_cache_path = None
        self.log_live_path = None
        self.log_bulk_path = None
        self.log_invalid_path = None

        # Data holders for loaded rules
        self.keyword_subject_to_delete1_keywords = set()
        self.trade_keywords = set()
        self.my_cliente_emails = set()
        self.dacs_notmine_emails = set()
        self.my_client_keywords = set()
        self.dacs_notmine_keywords = set()
        self.trade_details_emails = set()
        self.research_emails = set()
        self.research_keywords = set()
        self.boss_emails = set()
        self.keyword_subject_to_delete_keywords = set()

        self.smtp_cache = {}
        self.new_smtp_entries = {}
        self._smtp_cache_lock = threading.Lock()  # Guards smtp_cache and new_smtp_entries across threads

        self.live_running = False
        self.invalid_logger = None
        self.live_logger = None
        self.bulk_logger = None

        # Live mode scheduling state
        # Initialized to None to explicitly indicate no check has occurred yet
        self.last_midnight_check_hour = None

        try:
            self._load_config()
            self.setup_paths()
            self.setup_logging()

            self.load_data() # Load all data from Excel

            self.invalid_logger.info("EmailSorter initialized successfully.")
        except Exception as e:
            error_message = f"Initialization error: {e}"
            print(error_message)
            if self.invalid_logger:
                self.invalid_logger.error(f"InitializationError||EmailSorter.__init__|{error_message}")
            messagebox.showerror("Initialization Error", error_message)
            raise

    def _load_config(self):
        """
        Loads configuration from the JSON file.
        Ensures essential paths and sheet mappings are present and correctly formatted.
        """
        try:
            print(f"Attempting to load config from: {os.path.abspath(self.config_path)}")
            with open(self.config_path, 'r') as f:
                self.config = json.load(f)

            print(f"Keys found in sheet_map: {list(self.config['sheet_map'].keys())}")

            required_top_level_keys = ['xls_path', 'smtp_cache_path', 'log_live_path', 'log_bulk_path', 'log_invalid_path', 'smtp_fallback_dir', 'sheet_map']
            for key in required_top_level_keys:
                if key not in self.config:
                    raise ValueError(f"Missing required top-level configuration key: '{key}'")

            # Define expected structure for each rule type in sheet_map
            expected_rule_structure = {
                "KeywordSubject_ToDelete1": {"sheet": str, "columns": list, "match_field": str, "destination_name": str},
                "TradeKeyword": {"sheet": str, "column": str, "match_field": str, "destination_name": str},
                "MyClienteMailAddresses": {"sheet": str, "column": str, "destination_name": str},
                "DACSNotMineEmail": {"sheet": str, "column": str, "destination_name": str},
                "MyClientKeywords": {"sheet": str, "columns": list, "match_field": str, "destination_name": str},
                "DACSNotMineKeyword": {"sheet": str, "columns": list, "match_field": str, "destination_name": str},
                "TradeDetailseMailAddresses": {"sheet": str, "column": str, "destination_name": str},
                "ResearchEmail": {"sheet": str, "column": str, "destination_name": str},
                "ResearchKeyword": {"sheet": str, "column": str, "match_field": str, "destination_name": str},
                "BossEmail": {"sheet": str, "column": str, "destination_name": str},
                "KeywordSubject_ToDelete": {"sheet": str, "column": str, "match_field": str, "destination_name": str},
                "SMTPResolutionCache": {"sheet": str} # Column key is optional/can be null
            }

            for rule_name, required_keys in expected_rule_structure.items():
                if rule_name not in self.config['sheet_map']:
                    raise ValueError(f"Missing required rule configuration in sheet_map: '{rule_name}'")

                rule_config = self.config['sheet_map'][rule_name]
                for key, expected_type in required_keys.items():
                    if key not in rule_config:
                        # Allow 'column' to be missing if 'columns' is expected, or vice-versa
                        if (key == "column" and "columns" in rule_config) or \
                           (key == "columns" and "column" in rule_config):
                            continue # One of them is present, which is fine
                        if key == "match_field" and rule_name in ["MyClienteMailAddresses", "DACSNotMineEmail", 
                                                                  "TradeDetailseMailAddresses", "ResearchEmail", 
                                                                  "BossEmail"]:
                            continue # These don't need match_field as they're email address based
                        raise ValueError(f"Missing required key '{key}' for rule '{rule_name}' in sheet_map.")
                    if not isinstance(rule_config[key], expected_type):
                        # Special handling for column/columns
                        if (key == "columns" and not (isinstance(rule_config[key], list) and len(rule_config[key]) > 0)) or \
                           (key == "column" and not isinstance(rule_config[key], str)):
                            raise ValueError(f"Invalid type for key '{key}' in rule '{rule_name}'. Expected {expected_type.__name__}.")

                # Specific checks for match_field if present
                if "match_field" in rule_config and rule_config["match_field"] not in ['subject_only', 'subject_and_body']:
                    raise ValueError(f"Invalid 'match_field' for '{rule_name}'. Must be 'subject_only' or 'subject_and_body'.")

            print(f"Configuration loaded successfully from {self.config_path}")
        except FileNotFoundError:
            raise FileNotFoundError(f"Configuration file {self.config_path} not found.")
        except json.JSONDecodeError as e:
            raise ValueError(f"Invalid JSON in configuration file: {e}")
        except ValueError as e:
            raise e

    def setup_paths(self):
        """Sets up file paths from the loaded configuration."""
        self.xls_path         = self.config['xls_path']
        self.smtp_cache_path  = self.config['smtp_cache_path']
        self.smtp_fallback_dir = self.config['smtp_fallback_dir']
        self.log_live_path    = self.config['log_live_path']
        self.log_bulk_path    = self.config['log_bulk_path']
        self.log_invalid_path = self.config['log_invalid_path']

        # Ensure all log and cache directories exist before logging starts
        os.makedirs(os.path.dirname(self.log_live_path) or '.', exist_ok=True)
        os.makedirs(os.path.dirname(self.log_bulk_path) or '.', exist_ok=True)
        os.makedirs(os.path.dirname(self.log_invalid_path) or '.', exist_ok=True)
        os.makedirs(os.path.dirname(self.smtp_cache_path) or '.', exist_ok=True)

        print(f"Paths set: Excel='{self.xls_path}', SMTP_Cache='{self.smtp_cache_path}', LiveLog='{self.log_live_path}', BulkLog='{self.log_bulk_path}', InvalidLog='{self.log_invalid_path}'")

    def setup_logging(self):
        """Configures logging for live, bulk, and invalid/warning email entries."""
        # invalid_logger is set to WARNING so both .warning() and .error() calls are captured
        self.invalid_logger = self._create_logger('invalid_log', self.log_invalid_path, level=logging.WARNING)
        self.live_logger    = self._create_logger('live_log',    self.log_live_path,    level=logging.INFO)
        self.bulk_logger    = self._create_logger('bulk_log',    self.log_bulk_path,    level=logging.INFO)
        print("Logging setup complete.")

    def _create_logger(self, name, log_path, level=logging.INFO):
        """Helper to create and configure a logger."""
        logger = logging.getLogger(name)
        logger.setLevel(level)

        if logger.handlers:
            for handler in list(logger.handlers):
                logger.removeHandler(handler)

        handler = logging.FileHandler(log_path, mode='a', encoding='utf-8')
        formatter = logging.Formatter('%(asctime)s|%(levelname)s|%(message)s')
        handler.setFormatter(formatter)
        logger.addHandler(handler)
        return logger

    def _load_rule(self, attr_name, loader_fn, rule_name, description):
        """
        Helper used by load_data to load a single rule dataset.
        Sets self.<attr_name> to the loaded value, or leaves it at its default
        (empty set) and logs a warning if loading fails.
        """
        try:
            value = loader_fn(rule_name)
            setattr(self, attr_name, value)
            print(f"Loaded {len(value)} {description}.")
        except Exception as e:
            self.invalid_logger.error(f"DataLoadError|{rule_name}|load_data|{e}")
            print(f"Warning: Could not load {rule_name}: {e}")

    def load_data(self):
        """
        Loads all rule datasets (email addresses, keywords, SMTP cache)
        from the configured Excel file. Each rule loads independently so a
        single bad sheet does not abort the rest.
        """
        try:
            self.tables = pd.read_excel(self.xls_path, sheet_name=None, dtype=str, engine='openpyxl')
            print(f"Excel file '{self.xls_path}' loaded successfully.")
        except FileNotFoundError:
            error_msg = f"Excel file not found: {self.xls_path}"
            self.invalid_logger.error(f"FileLoadError||load_data|{error_msg}")
            raise FileNotFoundError(error_msg)
        except Exception as e:
            error_msg = f"Error reading Excel file: {e}"
            self.invalid_logger.error(f"ExcelReadError||load_data|{error_msg}")
            raise ValueError(error_msg)

        self._load_rule('keyword_subject_to_delete1_keywords', self._load_keywords,              'KeywordSubject_ToDelete1',    "'to delete 1' subject keywords")
        self._load_rule('trade_keywords',                      self._load_single_column_keywords, 'TradeKeyword',                'trade keywords')
        self._load_rule('my_cliente_emails',                   self._load_email_addresses,        'MyClienteMailAddresses',      'client email addresses')
        self._load_rule('dacs_notmine_emails',                 self._load_email_addresses,        'DACSNotMineEmail',            'non-mine email addresses')
        self._load_rule('my_client_keywords',                  self._load_keywords,               'MyClientKeywords',            'my client keywords')
        self._load_rule('dacs_notmine_keywords',               self._load_keywords,               'DACSNotMineKeyword',          'non-mine keywords')
        self._load_rule('trade_details_emails',                self._load_email_addresses,        'TradeDetailseMailAddresses',  'trade details email addresses')
        self._load_rule('research_emails',                     self._load_email_addresses,        'ResearchEmail',               'research email addresses')
        self._load_rule('research_keywords',                   self._load_single_column_keywords, 'ResearchKeyword',             'research keywords')
        self._load_rule('boss_emails',                         self._load_email_addresses,        'BossEmail',                   'boss email addresses')
        self._load_rule('keyword_subject_to_delete_keywords',  self._load_single_column_keywords, 'KeywordSubject_ToDelete',     "'to delete' subject keywords")

        try:
            self.smtp_cache = self._load_smtp_cache()
            print(f"Loaded {len(self.smtp_cache)} SMTP cache entries.")
        except Exception as e:
            self.invalid_logger.error(f"DataLoadError|SMTPResolutionCache|load_data|{e}")
            print(f"Warning: Could not load SMTPResolutionCache: {e}")

        self.new_smtp_entries = {}

    def _load_email_addresses(self, rule_name):
        """Loads email addresses from a specified Excel sheet and column."""
        sheet_config = self.config['sheet_map'][rule_name]
        sheet_name = sheet_config['sheet']
        column_name = sheet_config['column']

        if sheet_name not in self.tables:
            raise ValueError(f"Sheet '{sheet_name}' not found in Excel file for '{rule_name}'")

        df = self.tables[sheet_name]
        if column_name not in df.columns:
            raise ValueError(f"Column '{column_name}' not found in sheet '{sheet_name}' for '{rule_name}'")

        return set(df[column_name].dropna().astype(str).str.strip().str.lower())

    def _load_single_column_keywords(self, rule_name):
        """
        Loads keywords from a specified Excel sheet with a single column.
        Includes stripping leading/trailing whitespace.
        """
        sheet_config = self.config['sheet_map'][rule_name]
        sheet_name = sheet_config['sheet']
        column_name = sheet_config['column']

        if sheet_name not in self.tables:
            raise ValueError(f"Sheet '{sheet_name}' not found in Excel file for '{rule_name}'")

        df = self.tables[sheet_name]
        if column_name not in df.columns:
            raise ValueError(f"Column '{column_name}' not found in sheet '{sheet_name}' for '{rule_name}'")

        keywords = set(
            df[column_name]
            .dropna()
            .astype(str)
            .str.strip()
            .str.lower()
        )
        keywords.discard('')
        return keywords

    def _load_keywords(self, rule_name):
        """
        Loads keywords from a specified Excel sheet and multiple columns.
        Includes stripping leading/trailing whitespace.
        """
        sheet_config = self.config['sheet_map'][rule_name]
        sheet_name = sheet_config['sheet']
        columns = sheet_config['columns'] # Expects a list of columns

        if sheet_name not in self.tables:
            raise ValueError(f"Sheet '{sheet_name}' not found in Excel file for '{rule_name}'")

        df = self.tables[sheet_name]

        missing_columns = [col for col in columns if col not in df.columns]
        if missing_columns:
            raise ValueError(f"Columns {missing_columns} not found in sheet '{sheet_name}' for '{rule_name}'")

        keywords = set()
        for column in columns:
            column_values = df[column].dropna().astype(str).str.strip().str.lower()
            keywords.update(column_values)

        keywords.discard('')
        return keywords

    def _load_smtp_cache(self):
        """Loads SMTP resolution cache from the separate SMTP cache Excel file."""
        cache_sheet_name = self.config['sheet_map']['SMTPResolutionCache']['sheet']
        
        # Check if the SMTP cache file exists
        if not os.path.exists(self.smtp_cache_path):
            self.invalid_logger.warning(
                f"SMTP cache file '{self.smtp_cache_path}' not found. Starting with empty cache."
            )
            return {}
        
        try:
            # Use pandas to read just the cache sheet to avoid loading entire workbook
            cache_df = pd.read_excel(self.smtp_cache_path, sheet_name=cache_sheet_name, dtype=str, engine='openpyxl')
            
            if 'EntryName' in cache_df.columns and 'SMTPAddress' in cache_df.columns:
                return dict(zip(
                    cache_df['EntryName'].fillna('').astype(str).str.lower(),
                    cache_df['SMTPAddress'].fillna('').astype(str).str.lower()
                ))
            else:
                self.invalid_logger.warning(
                    f"Missing 'EntryName' or 'SMTPAddress' columns in '{cache_sheet_name}' sheet. "
                    "SMTP cache will not be loaded."
                )
                return {}
        except Exception as e:
            self.invalid_logger.error(f"SMTPCacheLoadError||_load_smtp_cache|Failed to load SMTP cache: {e}")
            return {}

    def get_smtp_address(self, outlook_namespace, entry):
        """
        Resolves the SMTP address for an Outlook recipient or sender entry.
        Resolution order: in-memory cache → MAPI property → Address field fallback.
        Newly resolved addresses are added to both caches under a lock since
        this method is called from the live-mode background thread.
        """
        if not entry:
            self.invalid_logger.warning("NullEntry||get_smtp_address|Received a None entry.")
            return None

        name    = getattr(entry, 'Name',    '') or ''
        address = getattr(entry, 'Address', '') or ''
        name_key = (name.lower() if name else address.lower()) or ''

        if not name_key:
            self.invalid_logger.warning(f"EmptyNameKey|Name: '{name}', Address: '{address}'|get_smtp_address|No usable identifier for SMTP lookup.")
            return None

        # Check in-memory cache first to avoid repeated MAPI calls
        cached = self.smtp_cache.get(name_key)
        if cached:
            return cached

        smtp = None
        try:
            # Primary: read the PR_SMTP_ADDRESS MAPI property
            smtp = entry.PropertyAccessor.GetProperty("http://schemas.microsoft.com/mapi/proptag/0x39FE001E")
            if smtp:
                smtp = smtp.lower()
        except Exception:
            pass

        # Fallback: use the Address field if MAPI property was unavailable
        if not smtp and address:
            smtp = address.lower()

        if smtp:
            # Lock because run_live runs on a background thread and the GUI
            # thread may call save_smtp_cache concurrently
            with self._smtp_cache_lock:
                self.new_smtp_entries[name_key] = smtp
                self.smtp_cache[name_key]       = smtp
            return smtp
        else:
            self.invalid_logger.info(f"NoSMTPResolution|Name: '{name}', Address: '{address}'|get_smtp_address|Could not resolve SMTP address.")
            return None

    def extract_addresses(self, outlook_namespace, mail):
        """Extracts all relevant email addresses (recipients and sender) from a mail item."""
        recipients = set()

        # Extract recipients (To, Cc, Bcc)
        try:
            for rec in mail.Recipients:
                smtp = self.get_smtp_address(outlook_namespace, rec)
                if smtp:
                    recipients.add(smtp)
        except Exception as e:
            self.invalid_logger.error(f"RecipientParseError|{mail.Subject or 'NoSubject'}|extract_addresses|{e}")

        # Extract sender
        sender = None
        try:
            sender = self.get_smtp_address(outlook_namespace, mail.Sender)
            if sender:
                recipients.add(sender)
        except Exception as e:
            self.invalid_logger.error(f"SenderParseError|{mail.Subject or 'NoSubject'}|extract_addresses|{e}")

        if not recipients:
            self.invalid_logger.error(f"NoAddressesFound|Subject: '{mail.Subject or 'NoSubject'}'|extract_addresses|No sender or recipient addresses extracted.")
        return recipients
    
    def extract_sender_address_only(self, outlook_namespace, mail):
        """
        Extracts only the email address from the "From" field (sender) of a mail item.
        Used specifically for ResearchEmail (Rule 8) and BossEmail (Rule 10) rules.
        """
        sender_address = None
        
        try:
            sender = mail.Sender
            if sender:
                sender_address = self.get_smtp_address(outlook_namespace, sender)
        except Exception as e:
            self.invalid_logger.error(f"SenderParseError|{mail.Subject or 'NoSubject'}|extract_sender_address_only|{e}")
            
        return {sender_address} if sender_address else set()

    def _strip_html_tags(self, html_string):
        """
        Strips HTML tags and normalises whitespace to produce clean plain text
        for keyword matching. Uses html.unescape() to handle all HTML entities
        rather than maintaining a manual replacement list.
        """
        if not html_string:
            return ""

        # Decode all HTML entities (&nbsp;, &amp;, numeric entities, etc.)
        import html
        html_string = html.unescape(html_string)

        # Remove script and style blocks entirely (content is not matchable text)
        clean_text = re.sub(r'<script[^>]*>.*?</script>', '', html_string, flags=re.DOTALL | re.IGNORECASE)
        clean_text = re.sub(r'<style[^>]*>.*?</style>',  '', clean_text,  flags=re.DOTALL | re.IGNORECASE)

        # Convert block-level tags to newlines to preserve paragraph boundaries
        clean_text = re.sub(r'<br\s*/?>', '\n',   clean_text, flags=re.IGNORECASE)
        clean_text = re.sub(r'</p>',      '\n\n', clean_text, flags=re.IGNORECASE)

        # Strip all remaining HTML tags
        clean_text = re.sub(r'<[^>]*>', '', clean_text)

        # Collapse all whitespace (spaces, tabs, newlines) to a single space
        clean_text = re.sub(r'\s+', ' ', clean_text).strip()

        return clean_text

    def keyword_match(self, mail, keywords, match_field="subject_and_body"):
        """
        Checks if any of the provided keywords matches a phrase/word in the
        subject or body of the email based on match_field, using regex for substring matching.
        It prioritizes the cleaned HTML body content if available.
        """
        try:
            subject = (mail.Subject or "").lower()

            # Get and clean HTML body first
            body_html_cleaned = ""
            try:
                if hasattr(mail, 'HTMLBody') and mail.HTMLBody:
                    body_html_cleaned = self._strip_html_tags(mail.HTMLBody).lower()
            except Exception as e:
                self.invalid_logger.warning(f"HTMLBodyReadError|{mail.Subject or 'NoSubject'}|keyword_match|Failed to read or strip HTMLBody: {e}")

            # Fallback to plain text body if HTML is empty or failed
            body_plain_text = (mail.Body or "").lower()

            target_content_strings = []
            if match_field == "subject_only":
                target_content_strings.append(subject)
            elif match_field == "subject_and_body":
                target_content_strings.append(subject)
                # Use cleaned HTML body if it has content, otherwise fall back to plain text body
                if body_html_cleaned:
                    target_content_strings.append(body_html_cleaned)
                else:
                    target_content_strings.append(body_plain_text)
            else:
                self.invalid_logger.error(f"InvalidMatchField|{mail.Subject or 'NoSubject'}|keyword_match|Unknown match_field: {match_field}. Defaulting to subject_and_body.")
                target_content_strings.append(subject)
                if body_html_cleaned:
                    target_content_strings.append(body_html_cleaned)
                else:
                    target_content_strings.append(body_plain_text)

            for keyword in keywords:
                # Escape the keyword to treat it as a literal string in regex
                # No word boundaries for substring matching
                pattern = re.escape(keyword)
                regex = re.compile(pattern, re.IGNORECASE)

                for content_string in target_content_strings:
                    if regex.search(content_string):
                        return keyword # Return the first matching keyword
            return None
        except Exception as e:
            self.invalid_logger.error(f"KeywordMatchError|{mail.Subject or 'NoSubject'}|keyword_match|{e}")
            return None

    def log_email(self, logger, outlook_namespace, mail, match_info, dest_folder_name):
        """Logs processed email information to the specified logger."""
        try:
            sent_on = mail.SentOn
            date_str = sent_on.strftime("%Y-%m-%d")
            time_str = sent_on.strftime("%H:%M:%S")

            sender_smtp = self.get_smtp_address(outlook_namespace, mail.Sender) or "Unknown"

            subject = (mail.Subject or "NoSubject").replace('|', ' ').replace('\n', ' ').strip()

            log_entry = f"{date_str}|{time_str}|{sender_smtp}|{subject}|{match_info}|{dest_folder_name}"
            logger.info(log_entry)
        except Exception as e:
            self.invalid_logger.error(
                f"LogFormatError|Subject: '{getattr(mail, 'Subject', 'NoSubject') or 'NoSubject'}'|"
                f"log_email|Failed to format log entry: {e}"
            )

    def _get_or_create_outlook_folder(self, outlook_namespace, folder_path):
        """
        Gets an Outlook folder object by its path, creating it and any necessary
        parent folders if they don't exist.
        Path can be nested, e.g., "Inbox\\SubFolder\\SubSubFolder".
        It handles paths starting with standard folder names or implicitly
        creates subfolders under Inbox if no root is specified.
        """
        path_parts = folder_path.split('\\')
        current_folder = None

        # Determine the initial root folder
        first_part_lower = path_parts[0].lower()

        # Try to get default folders by their common names first
        if first_part_lower == "inbox":
            current_folder = outlook_namespace.GetDefaultFolder(self._OL_FOLDER_INBOX) # olFolderInbox
            path_parts = path_parts[1:] # Remove "Inbox" from parts to process
        elif first_part_lower == "sent items":
            current_folder = outlook_namespace.GetDefaultFolder(self._OL_FOLDER_SENT) # olFolderSentMail
            path_parts = path_parts[1:] # Remove "Sent Items" from parts to process
        else:
            # If the path doesn't start with "Inbox" or "Sent Items", assume it's
            # intended as a subfolder of Inbox.
            current_folder = outlook_namespace.GetDefaultFolder(self._OL_FOLDER_INBOX) # olFolderInbox
            # path_parts remains as is, e.g., ["DACS-My"] will be created directly under Inbox

        # Iterate through the remaining parts of the path, creating folders as needed
        for sub_folder_name in path_parts:
            # Skip empty parts (e.g., if path was "Inbox\\")
            if not sub_folder_name:
                continue

            try:
                # Attempt to get the subfolder
                current_folder = current_folder.Folders.Item(sub_folder_name)
            except Exception:
                # Folder does not exist, create it
                current_folder = current_folder.Folders.Add(sub_folder_name)
                print(f"Created Outlook folder: '{sub_folder_name}' under '{current_folder.Parent.Name}'")
        return current_folder

    def is_mail_open_in_inspector(self, outlook_namespace, mail):
        """
        Returns True if this MailItem is currently open in any Inspector window.
        """
        try:
            application = outlook_namespace.Application
            inspectors = application.Inspectors

            for inspector in inspectors:
                # Compare the underlying COM object to see if it's the same item
                if inspector.CurrentItem and inspector.CurrentItem.EntryID == mail.EntryID:
                    return True

            return False
        except Exception:
            # Fail-safe: if we can't determine, treat as not open
            return False

    def process_email(self, outlook_namespace, mail, logger, folder_objects_map):
        """
        Processes a single email: finds matching rules and moves the email.
        Rule order updated according to sheet numbering in config_v38.10.json:
        1. KeywordSubject_ToDelete1
        2. TradeKeyword
        3. MyClienteMailAddresses (checks ALL fields: From, To, CC, BCC)
        4. DACSNotMineEmail (checks ALL fields: From, To, CC, BCC)
        5. MyClientKeywords
        6. DACSNotMineKeyword
        7. TradeDetailseMailAddresses (checks ALL fields: From, To, CC, BCC)
        8. ResearchEmail (checks ONLY "From" sender field)
        9. ResearchKeyword
        10. BossEmail (checks ONLY "From" sender field)
        11. KeywordSubject_ToDelete
        """
        try:
            # Skip emails currently open in an Inspector to avoid COM conflicts
            if self.is_mail_open_in_inspector(outlook_namespace, mail):
                self.log_email(
                    logger,
                    outlook_namespace,
                    mail,
                    "Skipped processing because item is open in Inspector",
                    mail.Parent.Name  # or any destination name you normally log in this case
                )
                return False
            
            recipients = self.extract_addresses(outlook_namespace, mail)

            # Rule 1: Keyword in subject ONLY from KeywordSubject_ToDelete1 (highest priority)
            if self.keyword_match(mail, self.keyword_subject_to_delete1_keywords, match_field="subject_only"):
                dest_folder_name = self.config['sheet_map']['KeywordSubject_ToDelete1']['destination_name']
                mail.Move(folder_objects_map[dest_folder_name])
                self.log_email(logger, outlook_namespace, mail, "Matched by KeywordSubject_ToDelete1", dest_folder_name)
                return True

            # Rule 2: Keyword in subject ONLY from TradeKeyword
            if self.keyword_match(mail, self.trade_keywords, match_field="subject_only"):
                dest_folder_name = self.config['sheet_map']['TradeKeyword']['destination_name']
                mail.Move(folder_objects_map[dest_folder_name])
                self.log_email(logger, outlook_namespace, mail, "Matched by TradeKeyword", dest_folder_name)
                return True

            # Rule 3: Email address in MyClienteMailAddresses (checks ALL fields)
            if any(addr in self.my_cliente_emails for addr in recipients):
                dest_folder_name = self.config['sheet_map']['MyClienteMailAddresses']['destination_name']
                mail.Move(folder_objects_map[dest_folder_name])
                self.log_email(logger, outlook_namespace, mail, "Matched by MyClienteMailAddresses", dest_folder_name)
                return True

            # Rule 4: Email address in DACSNotMineEmail (checks ALL fields)
            if any(addr in self.dacs_notmine_emails for addr in recipients):
                dest_folder_name = self.config['sheet_map']['DACSNotMineEmail']['destination_name']
                mail.Move(folder_objects_map[dest_folder_name])
                self.log_email(logger, outlook_namespace, mail, "Matched by DACSNotMineEmail", dest_folder_name)
                return True

            # Rule 5: Keyword in subject/body from MyClientKeywords
            if self.keyword_match(mail, self.my_client_keywords, match_field="subject_and_body"):
                dest_folder_name = self.config['sheet_map']['MyClientKeywords']['destination_name']
                mail.Move(folder_objects_map[dest_folder_name])
                self.log_email(logger, outlook_namespace, mail, f"Matched by MyClientKeywords", dest_folder_name)
                return True

            # Rule 6: Keyword in subject/body from DACSNotMineKeyword
            if self.keyword_match(mail, self.dacs_notmine_keywords, match_field="subject_and_body"):
                dest_folder_name = self.config['sheet_map']['DACSNotMineKeyword']['destination_name']
                mail.Move(folder_objects_map[dest_folder_name])
                self.log_email(logger, outlook_namespace, mail, f"Matched by DACSNotMineKeyword", dest_folder_name)
                return True

            # Rule 7: Email address in TradeDetailseMailAddresses (checks ALL fields)
            if any(addr in self.trade_details_emails for addr in recipients):
                dest_folder_name = self.config['sheet_map']['TradeDetailseMailAddresses']['destination_name']
                mail.Move(folder_objects_map[dest_folder_name])
                self.log_email(logger, outlook_namespace, mail, "Matched by TradeDetailseMailAddresses", dest_folder_name)
                return True

            # Rule 8: Email address in ResearchEmail (ONLY checks "From" sender field)
            sender_addresses = self.extract_sender_address_only(outlook_namespace, mail)
            if any(addr in self.research_emails for addr in sender_addresses):
                dest_folder_name = self.config['sheet_map']['ResearchEmail']['destination_name']
                mail.Move(folder_objects_map[dest_folder_name])
                self.log_email(logger, outlook_namespace, mail, "Matched by ResearchEmail (From field only)", dest_folder_name)
                return True

            # Rule 9: Keyword in subject ONLY from ResearchKeyword
            if self.keyword_match(mail, self.research_keywords, match_field="subject_only"):
                dest_folder_name = self.config['sheet_map']['ResearchKeyword']['destination_name']
                mail.Move(folder_objects_map[dest_folder_name])
                self.log_email(logger, outlook_namespace, mail, "Matched by ResearchKeyword", dest_folder_name)
                return True

            # Rule 10: Email address in BossEmail (ONLY checks "From" sender field)
            if any(addr in self.boss_emails for addr in sender_addresses):
                dest_folder_name = self.config['sheet_map']['BossEmail']['destination_name']
                mail.Move(folder_objects_map[dest_folder_name])
                self.log_email(logger, outlook_namespace, mail, "Matched by BossEmail (From field only)", dest_folder_name)
                return True

            # Rule 11: Keyword in subject ONLY from KeywordSubject_ToDelete (lowest priority)
            if self.keyword_match(mail, self.keyword_subject_to_delete_keywords, match_field="subject_only"):
                dest_folder_name = self.config['sheet_map']['KeywordSubject_ToDelete']['destination_name']
                mail.Move(folder_objects_map[dest_folder_name])
                self.log_email(logger, outlook_namespace, mail, f"Matched by KeywordSubject_ToDelete", dest_folder_name)
                return True

            # If no rules matched, log to original folder name
            self.log_email(logger, outlook_namespace, mail, "No matching rules", mail.Parent.Name)
            return False

        except Exception as e:
            subject = getattr(mail, 'Subject', 'Unknown') or 'NoSubject'
            self.invalid_logger.error(f"EmailProcessingError|Subject: '{subject}'|process_email|{e}")
            print(f"Error processing email '{subject}': {e}")
            return False

    def _get_live_mode_start_filter_time(self):
        """
        Determines the start_time_filter for live mode based on current day of week
        and an hourly midnight check.
        """
        now = datetime.datetime.now()
        today_midnight = now.replace(hour=0, minute=0, second=0, microsecond=0)

        # Calculate base start day for daily/weekly lookback
        base_start_day = today_midnight
        weekday = now.weekday() # Monday is 0, Sunday is 6

        if weekday == 0: # Monday, look back to last Friday
            base_start_day = today_midnight - datetime.timedelta(days=3)
        elif weekday == 5: # Saturday, look back to Friday
            base_start_day = today_midnight - datetime.timedelta(days=1)
        elif weekday == 6: # Sunday, look back to Friday
            base_start_day = today_midnight - datetime.timedelta(days=2)
        else: # Tue, Wed, Thu, Fri
            base_start_day = today_midnight - datetime.timedelta(days=1)

        # Flag to indicate if we should do a midnight check this iteration
        trigger_midnight_check = False

        if self.last_midnight_check_hour is None:
            # Always trigger midnight check on the very first run of live mode
            trigger_midnight_check = True
            print("Live mode: Initial (first run) midnight check triggered.")
        elif now.hour != self.last_midnight_check_hour:
            # If the hour has changed since the last midnight check, trigger one
            trigger_midnight_check = True
            print(f"Live mode: Hourly midnight check triggered for new hour {now.hour:02d}.")

        if trigger_midnight_check:
            start_time_filter = datetime.datetime.combine(base_start_day.date(), datetime.time.min)
            self.last_midnight_check_hour = now.hour # Update the last checked hour to the current hour
            print(f"    Processing from {start_time_filter.strftime('%Y-%m-%d %H:%M:%S')}.")
        else:
            # If not a midnight check, use the standard 5-minute lookback
            start_time_filter = now - datetime.timedelta(minutes=5)
            print(f"Live mode: Standard 5-minute lookback. Processing from {start_time_filter.strftime('%Y-%m-%d %H:%M:%S')}.")

        return start_time_filter

    def process_folder(self, outlook_namespace, folder_to_process, logger, date_filter_time, folder_objects_map):
        """
        Processes all emails within a given Outlook folder.
        Filters emails from the specified date/time onwards.
        Emails are processed in reverse order to handle deletion/moving without affecting iteration.
        This method takes outlook_namespace and target folder *objects* as arguments.
        `date_filter_time` is now a datetime.datetime object.
        """
        processed_count = 0
        folder_name = getattr(folder_to_process, 'Name', 'Unknown')
        print(f"Starting processing for folder: {folder_name}")

        try:
            if not hasattr(folder_to_process, 'Items'):
                self.invalid_logger.error(f"InvalidFolder|{folder_name}|process_folder|Folder object has no 'Items' attribute.")
                return 0

            messages = folder_to_process.Items
            messages.Sort("[ReceivedTime]", False) # Sort by ReceivedTime descending

            # Format the datetime object for Outlook's Restrict method
            date_filter_outlook_str = date_filter_time.strftime('%d/%m/%Y %H:%M %p')
            filter_string = f"[ReceivedTime] >= '{date_filter_outlook_str}'"
            print(f"Applying Outlook filter: {filter_string}")

            try:
                # Use ReceivedTime for filter as it's more reliable than SentOn for filtering when email arrived
                filtered_messages = messages.Restrict(filter_string)
                total_messages_to_process = filtered_messages.Count
                print(f"Found {total_messages_to_process} messages in {folder_name} matching date filter {date_filter_time}.")
            except Exception as restrict_error:
                self.invalid_logger.error(f"OutlookFilterError|{folder_name}|process_folder|Failed to apply filter '{filter_string}': {restrict_error}. Processing all messages and filtering manually.")
                print(f"Warning: Failed to apply Outlook filter: {restrict_error}. Processing all messages and filtering manually.")
                filtered_messages = messages # Fallback to all messages, then filter manually below
                total_messages_to_process = messages.Count # Initial count

            current_message_count = filtered_messages.Count

            for i in range(current_message_count, 0, -1):
                try:
                    mail = filtered_messages.Item(i)

                    # Ensure mail.ReceivedTime is timezone-naive for comparison to avoid issues
                    mail_received_time_naive = mail.ReceivedTime.replace(tzinfo=None)

                    # Manual datetime check for robustness if Outlook's Restrict failed or is imprecise
                    if mail_received_time_naive < date_filter_time:
                        continue # Skip emails older than the filter time

                    # Pass thread-local Outlook objects to process_email
                    if self.process_email(outlook_namespace, mail, logger, folder_objects_map):
                        processed_count += 1

                except Exception as msg_error:
                    subject = getattr(mail, 'Subject', 'Unknown') or 'NoSubject'
                    self.invalid_logger.error(f"MessageAccessError|Folder: '{folder_name}', Subject: '{subject}'|process_folder|{msg_error}")
                    print(f"Error accessing or processing message in {folder_name}: {msg_error}")
                    continue

        except Exception as e:
            self.invalid_logger.critical(f"FolderProcessingError|{folder_name}|process_folder|{e}")
            print(f"Critical error processing folder {folder_name}: {e}")
            return 0

        return processed_count

    def run_live(self):
        """
        Runs the email sorter in live mode, continuously monitoring and
        processing emails.
        For the first run, it checks from calculated start day (midnight) or last 5 mins.
        The frequency is 1 minute, with an hourly midnight reset.
        """
        self.live_running = True
        print("Starting live mode...")
        self.live_logger.info("Live mode started.")

        outlook_app = None
        outlook_namespace = None

        # Dictionary to hold all destination folder objects
        live_folder_objects = {}

        try:
            pythoncom.CoInitialize()
            outlook_app = win32com.client.Dispatch("Outlook.Application")
            outlook_namespace = outlook_app.GetNamespace("MAPI")

            # Initialize all required Outlook folder objects dynamically from config
            required_dest_paths = set()
            for rule_name, rule_config in self.config['sheet_map'].items():
                if 'destination_name' in rule_config:
                    required_dest_paths.add(rule_config['destination_name'])

            # Also ensure Inbox and Sent Items are available for processing and as potential destinations
            live_folder_objects["Inbox"] = outlook_namespace.GetDefaultFolder(self._OL_FOLDER_INBOX)
            live_folder_objects["Sent Items"] = outlook_namespace.GetDefaultFolder(self._OL_FOLDER_SENT)

            # Get or create all other specified destination folders
            for folder_path in required_dest_paths:
                # If a folder path is already a direct reference to Inbox or Sent Items, don't re-create.
                if folder_path.lower() != "inbox" and folder_path.lower() != "sent items":
                    live_folder_objects[folder_path] = self._get_or_create_outlook_folder(outlook_namespace, folder_path)

            print("Outlook initialized for live mode thread.")

            while self.live_running:
                # Determine the time filter based on the new scheduling logic
                start_time_filter = self._get_live_mode_start_filter_time()

                processed_inbox = 0
                processed_sent = 0

                try:
                    processed_inbox = self.process_folder(outlook_namespace, live_folder_objects["Inbox"], self.live_logger, start_time_filter, live_folder_objects)
                    processed_sent = self.process_folder(outlook_namespace, live_folder_objects["Sent Items"], self.live_logger, start_time_filter, live_folder_objects)
                except Exception as e:
                    self.invalid_logger.critical(f"LiveModeFatalError||run_live|A critical error occurred in live mode: {e}")
                    print(f"A critical error occurred in live mode: {e}. Stopping.")
                    self.live_running = False

                total_processed = processed_inbox + processed_sent
                if total_processed > 0:
                    print(f"Processed {total_processed} emails in live mode. Sleeping for 60 seconds...")
                else:
                    print("No new emails to process in live mode. Sleeping for 60 seconds...")

                # Sleep for 1 minute (60 seconds)
                for _ in range(60): # Loop for 60 seconds, checking stop_live flag
                    if not self.live_running:
                        print("Live mode stopped by user request during sleep.")
                        break
                    time.sleep(1)

        except Exception as e:
            self.invalid_logger.critical(f"LiveThreadSetupError||run_live|Failed to set up Outlook in live mode thread: {e}")
            print(f"Failed to set up Outlook in live mode thread: {e}. Live mode aborted.")
            self.live_running = False

        finally:
            self.live_logger.info("Live mode stopped.")
            print("Live mode gracefully stopped.")
            # Clean up COM objects for live mode
            if outlook_app:
                # Release specific folder objects first
                for folder_obj in live_folder_objects.values():
                    try: del folder_obj
                    except Exception as e: self.invalid_logger.warning(f"COMObjectCleanup|run_live|Failed to delete folder object: {e}")

                # Then release namespace and app
                try: del outlook_namespace
                except Exception as e: self.invalid_logger.warning(f"COMObjectCleanup|run_live|Failed to delete namespace object: {e}")
                try: del outlook_app
                except Exception as e: self.invalid_logger.warning(f"COMObjectCleanup|run_live|Failed to delete outlook app object: {e}")
            if 'pythoncom' in globals():
                try:
                    pythoncom.CoUninitialize()
                except Exception as e:
                    self.invalid_logger.error(f"COMUninitializeError|run_live|Failed to uninitialize COM: {e}")
                    print(f"Warning: Failed to uninitialize COM in live mode: {e}")

    def stop_live(self):
        """Signals the live mode to stop its execution loop."""
        self.live_running = False
        print("Stopping live mode...")

    def run_bulk(self, start_date, end_date): # Bulk mode now accepts start_date and end_date
        """Runs the email sorter in bulk mode for a specified date range."""
        print(f"Starting bulk processing for date range: {start_date} to {end_date}...")
        self.bulk_logger.info(f"Bulk mode started for date range: {start_date} to {end_date}.")

        outlook_app = None
        outlook_namespace = None

        # Dictionary to hold all destination folder objects for bulk mode
        bulk_folder_objects = {}

        processed_inbox = 0
        processed_sent = 0

        try:
            pythoncom.CoInitialize()
            outlook_app = win32com.client.Dispatch("Outlook.Application")
            outlook_namespace = outlook_app.GetNamespace("MAPI")

            # Initialize all required Outlook folder objects dynamically from config
            required_dest_paths = set()
            for rule_name, rule_config in self.config['sheet_map'].items():
                if 'destination_name' in rule_config:
                    required_dest_paths.add(rule_config['destination_name'])

            # Also ensure Inbox and Sent Items are available for processing and as potential destinations
            bulk_folder_objects["Inbox"] = outlook_namespace.GetDefaultFolder(self._OL_FOLDER_INBOX)
            bulk_folder_objects["Sent Items"] = outlook_namespace.GetDefaultFolder(self._OL_FOLDER_SENT)

            # Get or create all other specified destination folders
            for folder_path in required_dest_paths:
                if folder_path.lower() != "inbox" and folder_path.lower() != "sent items":
                    bulk_folder_objects[folder_path] = self._get_or_create_outlook_folder(outlook_namespace, folder_path)

            print("Outlook initialized for bulk mode thread.")

            # Convert selected dates (datetime.date) to datetime.datetime at midnight
            bulk_start_time_filter = datetime.datetime.combine(start_date, datetime.time.min)
            # End date filter needs to be up to the end of the day, so add almost one day
            bulk_end_time_filter = datetime.datetime.combine(end_date, datetime.time.max) # Max time for the end date

            # Process Inbox
            processed_inbox = self.process_folder_bulk(outlook_namespace, bulk_folder_objects["Inbox"], self.bulk_logger, bulk_start_time_filter, bulk_end_time_filter, bulk_folder_objects)
            # Process Sent Items
            processed_sent = self.process_folder_bulk(outlook_namespace, bulk_folder_objects["Sent Items"], self.bulk_logger, bulk_start_time_filter, bulk_end_time_filter, bulk_folder_objects)

        except Exception as e:
            self.invalid_logger.critical(f"BulkModeFatalError||run_bulk|A critical error occurred in bulk mode: {e}")
            print(f"A critical error occurred in bulk mode: {e}.")

        finally:
            total_processed = processed_inbox + processed_sent
            print(f"Bulk processing completed. Processed {total_processed} emails for {start_date} to {end_date}.")
            self.bulk_logger.info(f"Bulk mode completed for date range: {start_date} to {end_date}. Processed {total_processed} emails.")

            messagebox.showinfo("Bulk Processing Complete",
                                f"Processed {total_processed} emails for {start_date} to {end_date}\n"
                                f"Inbox: {processed_inbox} emails\n"
                                f"Sent Items: {processed_sent} emails")
            # Clean up COM objects for bulk mode
            if outlook_app:
                for folder_obj in bulk_folder_objects.values():
                    try: del folder_obj
                    except Exception as e: self.invalid_logger.warning(f"COMObjectCleanup|run_bulk|Failed to delete folder object: {e}")

                try: del outlook_namespace
                except Exception as e: self.invalid_logger.warning(f"COMObjectCleanup|run_bulk|Failed to delete namespace object: {e}")
                try: del outlook_app
                except Exception as e: self.invalid_logger.warning(f"COMObjectCleanup|run_bulk|Failed to delete outlook app object: {e}")
            if 'pythoncom' in globals():
                try:
                    pythoncom.CoUninitialize()
                except Exception as e:
                    self.invalid_logger.error(f"COMUninitializeError|run_bulk|Failed to uninitialize COM: {e}")
                    print(f"Warning: Failed to uninitialize COM in bulk mode: {e}")

    # New method for bulk mode folder processing with start and end dates
    def process_folder_bulk(self, outlook_namespace, folder_to_process, logger, start_date_time, end_date_time, folder_objects_map):
        """
        Processes all emails within a given Outlook folder for a specific date range.
        Emails are processed in reverse order to handle deletion/moving without affecting iteration.
        """
        processed_count = 0
        folder_name = getattr(folder_to_process, 'Name', 'Unknown')
        print(f"Starting bulk processing for folder: {folder_name} from {start_date_time} to {end_date_time}")

        try:
            if not hasattr(folder_to_process, 'Items'):
                self.invalid_logger.error(f"InvalidFolder|{folder_name}|process_folder_bulk|Folder object has no 'Items' attribute.")
                return 0

            messages = folder_to_process.Items
            messages.Sort("[ReceivedTime]", False) # Sort by ReceivedTime descending

            # Use both start and end date filters for Outlook's Restrict method
            start_date_outlook_str = start_date_time.strftime('%d/%m/%Y %H:%M %p')
            end_date_outlook_str = end_date_time.strftime('%d/%m/%Y %H:%M %p')

            filter_string = f"[ReceivedTime] >= '{start_date_outlook_str}' AND [ReceivedTime] <= '{end_date_outlook_str}'"
            print(f"Applying Outlook filter: {filter_string}")

            try:
                filtered_messages = messages.Restrict(filter_string)
                total_messages_to_process = filtered_messages.Count
                print(f"Found {total_messages_to_process} messages in {folder_name} matching date range filter.")
            except Exception as restrict_error:
                self.invalid_logger.error(f"OutlookFilterError|{folder_name}|process_folder_bulk|Failed to apply filter '{filter_string}': {restrict_error}. Processing all messages and filtering manually.")
                print(f"Warning: Failed to apply Outlook filter: {restrict_error}. Processing all messages and filtering manually.")
                filtered_messages = messages # Fallback to all messages, then filter manually below
                total_messages_to_process = messages.Count # Initial count

            current_message_count = filtered_messages.Count

            for i in range(current_message_count, 0, -1):
                try:
                    mail = filtered_messages.Item(i)

                    mail_received_time_naive = mail.ReceivedTime.replace(tzinfo=None)

                    # Manual datetime check for robustness if Outlook's Restrict failed or is imprecise
                    if not (start_date_time <= mail_received_time_naive <= end_date_time):
                        continue # Skip emails outside the desired range

                    if self.process_email(outlook_namespace, mail, logger, folder_objects_map):
                        processed_count += 1

                except Exception as msg_error:
                    subject = getattr(mail, 'Subject', 'Unknown') or 'NoSubject'
                    self.invalid_logger.error(f"MessageAccessError|Folder: '{folder_name}', Subject: '{subject}'|process_folder_bulk|{msg_error}")
                    print(f"Error accessing or processing message in {folder_name}: {msg_error}")
                    continue

        except Exception as e:
            self.invalid_logger.critical(f"FolderProcessingError|{folder_name}|process_folder_bulk|{e}")
            print(f"Critical error processing folder {folder_name}: {e}")
            return 0

        return processed_count
    
    
    @staticmethod
    def _ensure_label_relationship(xml_text: str) -> str:
        """
        Ensures the MSIP classification relationship exists in _rels/.rels.

        openpyxl silently drops this entry on every save.  Without it, Office
        ignores LabelInfo.xml even when the file is physically present in the
        zip — the label appears to be there but is never recognised.

        If the relationship is absent it is inserted before </Relationships>
        using a safe generated Id that does not collide with existing rIds.
        """
        if EmailSorter.LABEL_REL_TYPE in xml_text:
            return xml_text  # already present

        existing_ids = [int(n) for n in re.findall(r'Id="rId(\d+)"', xml_text)]
        next_id = max(existing_ids, default=0) + 1

        rel_entry = (
            f'<Relationship Id="rId{next_id}"'
            f' Type="{EmailSorter.LABEL_REL_TYPE}"'
            f' Target="{EmailSorter.LABEL_REL_TARGET}"/>'
        )
        return xml_text.replace("</Relationships>", f"{rel_entry}</Relationships>")

    @staticmethod
    def _inject_msip_label(src_path: str, dst_path: str) -> None:
        """
        Copies src_path to dst_path as a zip, ensuring:
          - docMetadata/LabelInfo.xml is present with the correct label XML
          - [Content_Types].xml has the Override entry for the label part
          - _rels/.rels has the classification relationship entry (root cause:
            openpyxl drops this on every save, making Office ignore the label)

        src_path and dst_path must be different paths.
        """
        label_bytes = EmailSorter.LABEL_INFO_XML.encode("utf-8")

        buf = BytesIO()
        with zipfile.ZipFile(src_path, "r") as zin, \
             zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zout:

            existing_names = {i.filename for i in zin.infolist()}

            for item in zin.infolist():
                data = zin.read(item.filename)

                if item.filename == EmailSorter.LABEL_PART_PATH:
                    # Replace whatever is there with our canonical label XML
                    zout.writestr(item, label_bytes)

                elif item.filename == EmailSorter.CONTENT_TYPES_PATH:
                    # Ensure the Override entry exists; add it if openpyxl dropped it
                    xml_text = data.decode("utf-8")
                    if EmailSorter.LABEL_PART_PATH not in xml_text:
                        override_entry = (
                            f'<Override PartName="/{EmailSorter.LABEL_PART_PATH}"'
                            f' ContentType="{EmailSorter.LABEL_CONTENT_TYPE}"/>'
                        )
                        xml_text = xml_text.replace("</Types>", f"{override_entry}</Types>")
                        data = xml_text.encode("utf-8")
                    zout.writestr(item, data)

                elif item.filename == EmailSorter.RELS_PATH:
                    # Restore the classification relationship dropped by openpyxl
                    xml_text = EmailSorter._ensure_label_relationship(data.decode("utf-8"))
                    zout.writestr(item, xml_text.encode("utf-8"))

                else:
                    zout.writestr(item, data)

            # If the label part was absent entirely, add it now
            if EmailSorter.LABEL_PART_PATH not in existing_names:
                zout.writestr(EmailSorter.LABEL_PART_PATH, label_bytes)

        buf.seek(0)
        with open(dst_path, "wb") as f:
            f.write(buf.read())


    def save_smtp_cache(self) -> None:
        """
        Saves newly resolved SMTP entries using openpyxl (no COM / xlwings),
        then re-injects the MSIP sensitivity label directly into the xlsx zip.

        Eliminates the win32com gen_py cache corruption problem entirely while
        preserving the Microsoft Information Protection label on every save.

        After appending rows, the SMTPResolutionCache Excel Table is resized to
        cover all data rows so new entries are included in the table range
        (auto-filter, formatting, structured references all stay correct).

        Falls back to a TXT file if Excel I/O fails and the user cancels retry.
        """
        # Take a snapshot under the lock so the live thread can keep resolving
        # addresses while we do the (potentially slow) file I/O
        with self._smtp_cache_lock:
            if not self.new_smtp_entries:
                print("No new SMTP entries to save.")
                return
            entries_to_save = dict(self.new_smtp_entries)

        cache_sheet_name  = self.config["sheet_map"]["SMTPResolutionCache"]["sheet"]
        txt_fallback_path = os.path.join(self.smtp_fallback_dir, "SMTP_cache_fallback.txt")
        tmp_path          = self.smtp_cache_path + ".tmp"
        tmp_path2         = self.smtp_cache_path + ".tmp2"

        saved_successfully = False

        while not saved_successfully:
            try:
                # ── 1. Load existing workbook ────────────────────────────────────
                wb = openpyxl.load_workbook(self.smtp_cache_path, keep_vba=True)
                ws = wb[cache_sheet_name]

                # ── 2. Build set of already-cached entry names ───────────────────
                existing_entries = {
                    str(ws.cell(row=r, column=1).value).strip().lower()
                    for r in range(2, ws.max_row + 1)
                    if ws.cell(row=r, column=1).value is not None
                }

                # ── 3. Append new entries ────────────────────────────────────────
                next_row      = ws.max_row + 1
                entries_added = 0

                for entry_name, smtp_address in entries_to_save.items():
                    if entry_name.strip().lower() not in existing_entries:
                        ws.cell(row=next_row, column=1, value=entry_name)
                        ws.cell(row=next_row, column=2, value=smtp_address)
                        existing_entries.add(entry_name.strip().lower())
                        next_row      += 1
                        entries_added += 1
                        print(f"Adding new cache entry: {entry_name} -> {smtp_address}")

                if entries_added == 0:
                    print("No new SMTP entries to add (all already cached).")
                    return

                # ── 4. Resize the SMTPResolutionCache table ───────────────────────
                # openpyxl writes data outside the table range when rows are appended;
                # update the table ref so it covers all rows including the new ones.
                new_last_row = next_row - 1  # last row with data after appending
                table_resized = False
                for tbl in ws.tables.values():
                    if tbl.displayName == cache_sheet_name or \
                       tbl.displayName.lower() == cache_sheet_name.lower():
                        from openpyxl.utils import get_column_letter
                        min_col = tbl.ref.split(":")[0]   # e.g. "A1"  → keep as-is
                        # Derive the max column letter from current ref end cell
                        old_end = tbl.ref.split(":")[1]   # e.g. "B50"
                        max_col_letter = ''.join(filter(str.isalpha, old_end))  # "B"
                        tbl.ref = f"{min_col}:{max_col_letter}{new_last_row}"
                        print(f"Resized table '{tbl.displayName}' to {tbl.ref}")
                        table_resized = True
                        break

                if not table_resized:
                    # Log a warning but don't fail — data is still written correctly
                    self.invalid_logger.warning(
                        f"TableResizeSkipped|save_smtp_cache|"
                        f"Table '{cache_sheet_name}' not found in sheet; "
                        "new rows written outside table range."
                    )
                    print(f"Warning: Table '{cache_sheet_name}' not found; rows appended outside table range.")

                # ── 5. Save to temp file via openpyxl ────────────────────────────
                wb.save(tmp_path)

                # ── 6. Re-inject MSIP label into a second temp, then atomic move ──
                # src and dst must differ — _inject_msip_label reads src while
                # writing dst; passing the same path truncates the file mid-read.
                self._inject_msip_label(tmp_path, tmp_path2)
                os.remove(tmp_path)                           # discard unlabelled temp
                shutil.move(tmp_path2, self.smtp_cache_path) # atomic replace of original

                time.sleep(1)  # brief pause to ensure the file handle is fully released

                self.invalid_logger.info(
                    f"SMTPCacheSave|save_smtp_cache|"
                    f"Saved {entries_added} new entries with MSIP label preserved."
                )
                print(f"Saved {entries_added} new SMTP entries with MSIP label preserved.")

                # Clear only the entries we successfully saved
                with self._smtp_cache_lock:
                    for key in entries_to_save:
                        self.new_smtp_entries.pop(key, None)
                saved_successfully = True

            except Exception as e:
                # Clean up both temp files if present
                for p in (tmp_path, tmp_path2):
                    if os.path.exists(p):
                        try:
                            os.remove(p)
                        except Exception:
                            pass

                error_msg = f"Error saving SMTP cache: {e}"
                print(error_msg)
                self.invalid_logger.error(
                    f"CacheSaveError||save_smtp_cache|{error_msg}"
                )

                retry_choice = messagebox.askretrycancel(
                    "Cache Save Error",
                    f"Failed to save SMTP cache:\n{e}\n\n"
                    "Click Retry to try again, or Cancel for TXT fallback."
                )

                if not retry_choice:
                    txt_choice = messagebox.askyesno(
                        "Save Cache as TXT",
                        f"Save new entries as TXT to:\n{self.smtp_fallback_dir}\n\n"
                        "Each line: EntryName<TAB>SMTPAddress"
                    )
                    if txt_choice:
                        try:
                            os.makedirs(self.smtp_fallback_dir, exist_ok=True)
                            with open(txt_fallback_path, "a", encoding="utf-8") as f:
                                for entry_name, smtp_address in entries_to_save.items():
                                    f.write(f"{entry_name}\t{smtp_address}\n")
                            self.invalid_logger.info(
                                f"SMTPCacheSaveFallback|save_smtp_cache|"
                                f"Saved {len(entries_to_save)} entries to TXT: {txt_fallback_path}"
                            )
                            print(
                                f"Saved {len(entries_to_save)} SMTP entries "
                                f"to TXT fallback: {txt_fallback_path}"
                            )
                            with self._smtp_cache_lock:
                                for key in entries_to_save:
                                    self.new_smtp_entries.pop(key, None)
                        except Exception as txt_e:
                            self.invalid_logger.error(
                                f"CacheSaveFallbackError||save_smtp_cache|{txt_e}"
                            )
                    break   # exit retry loop regardless of TXT outcome

                time.sleep(1)  # brief pause before manual retry
                
    def start_gui(self):
        """Starts the main Tkinter GUI for the Email Sorter."""
        root = tk.Tk()
        root.title("Email Sorter v38.18b")
        root.geometry("350x450")
        root.resizable(False, False)
        root.attributes('-topmost', True)

        header_label = tk.Label(root, text="Email Sorter", font=("Arial", 16, "bold"), fg="#333333")
        header_label.pack(pady=15)

        info_label = tk.Label(root, text="Choose operation mode:", font=("Arial", 10), fg="#555555")
        info_label.pack(pady=5)

        def pick_bulk():
            """Handles bulk mode selection, prompting for a date range."""
            cal_win = tk.Toplevel(root)
            cal_win.title("Select Date Range for Bulk Processing")
            cal_win.geometry("350x550")
            cal_win.resizable(False, False)
            cal_win.attributes('-topmost', True)
            cal_win.grab_set()

            # Start Date selection
            start_date_label = tk.Label(cal_win, text="Select Start Date:", font=("Arial", 10))
            start_date_label.pack(pady=(10, 0))
            cal_start = Calendar(cal_win, selectmode='day', date_pattern='yyyy-mm-dd',
                                background="blue", foreground="white",
                                headersbackground="blue", headersforeground="white",
                                selectbackground="green", selectforeground="white",
                                normalbackground="lightgray", weekendbackground="darkgray")
            cal_start.pack(pady=(0, 10))

            # End Date selection
            end_date_label = tk.Label(cal_win, text="Select End Date:", font=("Arial", 10))
            end_date_label.pack(pady=(10, 0))

            # Checkbox for "End Date as Today"
            end_date_today_var = tk.BooleanVar(value=True)
            end_date_checkbox = tk.Checkbutton(cal_win, text="End Date as Today", variable=end_date_today_var,
                                               font=("Arial", 9))
            end_date_checkbox.pack(anchor=tk.W, padx=10)

            cal_end = Calendar(cal_win, selectmode='day', date_pattern='yyyy-mm-dd',
                              background="blue", foreground="white",
                              headersbackground="blue", headersforeground="white",
                              selectbackground="green", selectforeground="white",
                              normalbackground="lightgray", weekendbackground="darkgray")
            cal_end.pack(pady=(0, 10))

            # Function to toggle end date calendar state based on checkbox
            def toggle_end_date_calendar():
                if end_date_today_var.get():
                    cal_end.config(state='disabled')
                else:
                    cal_end.config(state='normal')
            end_date_today_var.trace_add("write", lambda *args: toggle_end_date_calendar())
            toggle_end_date_calendar()

            def validate_dates_and_process():
                selected_start_date = cal_start.selection_get()

                if end_date_today_var.get():
                    selected_end_date = datetime.date.today()
                else:
                    selected_end_date = cal_end.selection_get()

                today = datetime.date.today()

                if selected_start_date > today:
                    messagebox.showerror("Invalid Date", "Start Date cannot be in the future.")
                    return
                if selected_end_date > today:
                    messagebox.showerror("Invalid Date", "End Date cannot be in the future.")
                    return
                if selected_start_date > selected_end_date:
                    messagebox.showerror("Invalid Date Range", "Start Date cannot be after End Date.")
                    return

                cal_win.destroy()
                root.destroy()
                threading.Thread(target=lambda: self.run_bulk(selected_start_date, selected_end_date), daemon=True).start()

            button_frame = tk.Frame(cal_win)
            button_frame.pack(pady=10)

            tk.Button(button_frame, text="Process", command=validate_dates_and_process,
                      bg="#28a745", fg="white", width=12, height=1,
                      font=("Arial", 10, "bold"), relief=tk.RAISED).pack(side=tk.LEFT, padx=10)
            tk.Button(button_frame, text="Cancel", command=cal_win.destroy,
                      bg="#dc3545", fg="white", width=12, height=1,
                      font=("Arial", 10, "bold"), relief=tk.RAISED).pack(side=tk.LEFT, padx=10)

        def pick_live():
            """Handles live mode selection, starting continuous monitoring."""
            root.destroy()

            live_win = tk.Tk()
            live_win.title("Live Mode - Email Sorter")
            live_win.geometry("320x160")
            live_win.resizable(False, False)
            live_win.attributes('-topmost', True)

            status_label = tk.Label(live_win, text="Live monitoring active...",
                                     font=("Arial", 12, "bold"), fg="green")
            status_label.pack(pady=20)

            info_label = tk.Label(live_win, text="Emails are being processed automatically in the background.",
                                   font=("Arial", 9), fg="#666666")
            info_label.pack(pady=5)

            def stop_and_close():
                self.stop_live()
                # Calling save_smtp_cache, which now handles the prompt
                self.save_smtp_cache()
                live_win.destroy()

            tk.Button(live_win, text="Stop Live Mode", command=stop_and_close,
                      bg="#dc3545", fg="white", width=18, height=1,
                      font=("Arial", 10, "bold"), relief=tk.RAISED).pack(pady=20)

            threading.Thread(target=self.run_live, daemon=True).start()

            live_win.protocol("WM_DELETE_WINDOW", stop_and_close)
            live_win.mainloop()

        button_frame = tk.Frame(root)
        button_frame.pack(pady=20)

        tk.Button(button_frame, text="Run Live Mode", command=pick_live,
                  bg="#007bff", fg="white", width=18, height=2,
                  font=("Arial", 11, "bold"), relief=tk.RAISED).pack(pady=8)
        tk.Button(button_frame, text="Run Bulk Mode", command=pick_bulk,
                  bg="#ffc107", fg="white", width=18, height=2,
                  font=("Arial", 11, "bold"), relief=tk.RAISED).pack(pady=8)

        footer_label = tk.Label(root, text="Live: Monitors emails based on smart schedule | Bulk: Process selected date range",
                                 font=("Arial", 8), fg="gray")
        footer_label.pack(side=tk.BOTTOM, pady=10)

        def on_closing():
            if messagebox.askyesno("Exit", "Do you want to save the SMTP cache before exiting? (Recommended)"):
                self.save_smtp_cache()
            root.destroy()

        root.protocol("WM_DELETE_WINDOW", on_closing)
        root.mainloop()

def main():
    """Main function to run the Email Sorter application."""
    sorter = None
    try:
        sorter = EmailSorter()
        sorter.start_gui()
    except Exception as e:
        print(f"Error starting Email Sorter application: {e}")
        if sorter and sorter.invalid_logger:
            sorter.invalid_logger.critical(f"AppStartupError||main|{e}")

if __name__ == "__main__":
    main()
