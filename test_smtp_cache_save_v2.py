"""
Simple test: open testsmtp.xlsm, append one test record, inject MSIP label.
Check the file manually afterwards.
"""

import os
import re
import zipfile
import shutil
from io import BytesIO
import openpyxl

# ── Config ───────────────────────────────────────────────────────────────────
FILE_PATH        = r"C:\Users\abc\Downloads\testsmtp.xlsm"
SHEET_NAME       = "SMTPResolutionCache"
TEST_ENTRY_NAME  = "Test User"
TEST_SMTP        = "testuser@example.com"

# ── MSIP label constants (from docMetadata/LabelInfo.xml) ────────────────────
LABEL_INFO_XML = (
    '<?xml version="1.0" encoding="utf-8" standalone="yes"?>'
    '<clbl:labelList xmlns:clbl="http://schemas.microsoft.com/office/2020/mipLabelMetadata">'
    '<clbl:label'
    ' id="{840e60c6-cef6-4cc0-a98d-364c7249d74b}"'
    ' enabled="1"'
    ' method="Privileged"'
    ' siteId="{b44900f1-2def-4c3b-9ec6-9020d604e19e}"'
    ' removed="0"'
    ' />'
    '</clbl:labelList>'
)
LABEL_PART_PATH    = "docMetadata/LabelInfo.xml"
LABEL_CONTENT_TYPE = "application/vnd.ms-office.classificationlabels+xml"
CONTENT_TYPES_PATH = "[Content_Types].xml"
RELS_PATH          = "_rels/.rels"
LABEL_REL_TYPE     = "http://schemas.microsoft.com/office/2020/02/relationships/classificationlabels"
LABEL_REL_TARGET   = "docMetadata/LabelInfo.xml"

# ── Temp file paths ───────────────────────────────────────────────────────────
tmp_path  = FILE_PATH + ".tmp"
tmp_path2 = FILE_PATH + ".tmp2"


def _ensure_label_relationship(xml_text: str) -> str:
    """
    Ensures the MSIP classification relationship exists in _rels/.rels.
    If absent, inserts it before </Relationships> with a safe generated Id.
    """
    if LABEL_REL_TYPE in xml_text:
        return xml_text  # already present

    # Find the highest existing rId number so we don't collide
    existing_ids = [int(n) for n in re.findall(r'Id="rId(\d+)"', xml_text)]
    next_id = max(existing_ids, default=0) + 1

    rel_entry = (
        f'<Relationship Id="rId{next_id}"'
        f' Type="{LABEL_REL_TYPE}"'
        f' Target="{LABEL_REL_TARGET}"/>'
    )
    return xml_text.replace("</Relationships>", f"{rel_entry}</Relationships>")


def inject_msip_label(src_path, dst_path):
    """Copies src_path to dst_path ensuring MSIP label part is present."""
    label_bytes = LABEL_INFO_XML.encode("utf-8")
    buf = BytesIO()

    with zipfile.ZipFile(src_path, "r") as zin, \
         zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zout:

        existing = {i.filename for i in zin.infolist()}

        for item in zin.infolist():
            data = zin.read(item.filename)

            if item.filename == LABEL_PART_PATH:
                # Replace existing label with canonical version
                zout.writestr(item, label_bytes)

            elif item.filename == CONTENT_TYPES_PATH:
                # Ensure the Override entry for the label part is registered
                xml_text = data.decode("utf-8")
                if LABEL_PART_PATH not in xml_text:
                    override_entry = (
                        f'<Override PartName="/{LABEL_PART_PATH}"'
                        f' ContentType="{LABEL_CONTENT_TYPE}"/>'
                    )
                    xml_text = xml_text.replace("</Types>", f"{override_entry}</Types>")
                    data = xml_text.encode("utf-8")
                zout.writestr(item, data)

            elif item.filename == RELS_PATH:
                # Ensure the classification relationship is registered
                xml_text = data.decode("utf-8")
                xml_text = _ensure_label_relationship(xml_text)
                zout.writestr(item, xml_text.encode("utf-8"))

            else:
                zout.writestr(item, data)

        # Add label part if it was absent entirely
        if LABEL_PART_PATH not in existing:
            zout.writestr(LABEL_PART_PATH, label_bytes)

    buf.seek(0)
    with open(dst_path, "wb") as f:
        f.write(buf.read())


try:
    print(f"Opening: {FILE_PATH}")
    wb = openpyxl.load_workbook(FILE_PATH, keep_vba=True)
    ws = wb[SHEET_NAME]

    next_row = ws.max_row + 1
    ws.cell(row=next_row, column=1, value=TEST_ENTRY_NAME)
    ws.cell(row=next_row, column=2, value=TEST_SMTP)
    print(f"Appended row {next_row}: '{TEST_ENTRY_NAME}' -> '{TEST_SMTP}'")

    wb.save(tmp_path)
    print(f"Saved to temp file: {tmp_path}")

    inject_msip_label(tmp_path, tmp_path2)
    os.remove(tmp_path)
    shutil.move(tmp_path2, FILE_PATH)
    print(f"MSIP label injected. File written to: {FILE_PATH}")
    print("Done — open the file in Excel and check the record and sensitivity label.")

except Exception as e:
    print(f"ERROR: {e}")
    for p in (tmp_path, tmp_path2):
        if os.path.exists(p):
            try:
                os.remove(p)
            except Exception:
                pass
