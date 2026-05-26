import os
import time
import zipfile
import shutil
from io import BytesIO
from tkinter import messagebox

import openpyxl


# ── MSIP label constants (extracted from your LabelInfo.xml) ────────────────
_LABEL_INFO_XML = (
    '<?xml version="1.0" encoding="utf-8" standalone="yes"?>'
    '<clbl:labelList xmlns:clbl="http://schemas.microsoft.com/office/2020/mipLabelMetadata">'
    '<clbl:label'
    ' id="{840e60c6-cef6-4cc0-a98d-364c7249d74b}"'
    ' enabled="1"'
    ' method="Privileged"'
    ' siteId="{b44900f1-2def-4c3b-9ec6-9020d604e19e}"'
    ' removed="0"'
    ' />'
    '</clbl:labelList>'
)

# Part path and content-type exactly as found in your [Content_Types].xml
_LABEL_PART_PATH   = "docMetadata/LabelInfo.xml"
_LABEL_CONTENT_TYPE = "application/vnd.ms-office.classificationlabels+xml"
_CONTENT_TYPES_PATH = "[Content_Types].xml"


def _inject_msip_label(src_path: str, dst_path: str) -> None:
    """
    Copies src_path to dst_path as a zip, ensuring:
      - docMetadata/LabelInfo.xml is present with the correct label XML
      - [Content_Types].xml has the Override entry for the label part

    Both src and dst may be the same path (in-place update via temp file).
    """
    label_bytes = _LABEL_INFO_XML.encode("utf-8")

    buf = BytesIO()
    with zipfile.ZipFile(src_path, "r") as zin, \
         zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zout:

        for item in zin.infolist():
            data = zin.read(item.filename)

            if item.filename == _LABEL_PART_PATH:
                # Replace whatever is there with our canonical label XML
                zout.writestr(item, label_bytes)

            elif item.filename == _CONTENT_TYPES_PATH:
                # Ensure the Override entry exists; add it if openpyxl dropped it
                xml_text = data.decode("utf-8")
                if _LABEL_PART_PATH not in xml_text:
                    override_entry = (
                        f'<Override PartName="/{_LABEL_PART_PATH}"'
                        f' ContentType="{_LABEL_CONTENT_TYPE}"/>'
                    )
                    # Insert before closing </Types>
                    xml_text = xml_text.replace("</Types>", f"{override_entry}</Types>")
                    data = xml_text.encode("utf-8")
                zout.writestr(item, data)

            else:
                zout.writestr(item, data)

        # If the label part was absent entirely, add it now
        existing_names = {i.filename for i in zin.infolist()}
        if _LABEL_PART_PATH not in existing_names:
            zout.writestr(_LABEL_PART_PATH, label_bytes)

    buf.seek(0)
    with open(dst_path, "wb") as f:
        f.write(buf.read())


def save_smtp_cache(self) -> None:
    """
    Saves newly resolved SMTP entries using openpyxl (no COM / xlwings),
    then re-injects the MSIP sensitivity label directly into the xlsx zip.

    Eliminates the win32com gen_py cache corruption problem entirely while
    preserving the Microsoft Information Protection label on every save.

    Falls back to a TXT file if Excel I/O fails and the user cancels retry.
    """
    if not self.new_smtp_entries:
        print("No new SMTP entries to save.")
        return

    cache_sheet_name  = self.config["sheet_map"]["SMTPResolutionCache"]["sheet"]
    txt_fallback_dir  = r"C:\Users\abc\Downloads"
    txt_fallback_path = os.path.join(txt_fallback_dir, "SMTP_cache_fallback.txt")
    tmp_path          = self.smtp_cache_path + ".tmp"

    saved_successfully = False

    while not saved_successfully:
        try:
            # ── 1. Load existing workbook ────────────────────────────────────
            wb = openpyxl.load_workbook(self.smtp_cache_path)
            ws = wb[cache_sheet_name]

            # ── 2. Build set of already-cached entry names ───────────────────
            existing_entries = {
                str(ws.cell(row=r, column=1).value).strip().lower()
                for r in range(2, ws.max_row + 1)
                if ws.cell(row=r, column=1).value is not None
            }

            # ── 3. Append new entries ────────────────────────────────────────
            next_row      = ws.max_row + 1
            entries_added = 0

            for entry_name, smtp_address in self.new_smtp_entries.items():
                if entry_name.strip().lower() not in existing_entries:
                    ws.cell(row=next_row, column=1, value=entry_name)
                    ws.cell(row=next_row, column=2, value=smtp_address)
                    existing_entries.add(entry_name.strip().lower())
                    next_row      += 1
                    entries_added += 1
                    print(f"Adding new cache entry: {entry_name} -> {smtp_address}")

            if entries_added == 0:
                print("No new SMTP entries to add (all already cached).")
                return

            # ── 4. Save to temp file via openpyxl ────────────────────────────
            wb.save(tmp_path)

            # ── 5. Re-inject MSIP label and write final file atomically ───────
            _inject_msip_label(tmp_path, tmp_path)        # label injected in-place
            shutil.move(tmp_path, self.smtp_cache_path)   # atomic replace

            time.sleep(1)  # ensure filesystem write completes before next access

            self.invalid_logger.info(
                f"SMTPCacheSave|save_smtp_cache|"
                f"Saved {entries_added} new entries with MSIP label preserved."
            )
            print(f"Saved {entries_added} new SMTP entries with MSIP label preserved.")

            self.new_smtp_entries.clear()
            saved_successfully = True

        except Exception as e:
            # Clean up temp file if it exists
            if os.path.exists(tmp_path):
                try:
                    os.remove(tmp_path)
                except Exception:
                    pass

            error_msg = f"Error saving SMTP cache: {e}"
            print(error_msg)
            self.invalid_logger.error(
                f"CacheSaveError||save_smtp_cache|{error_msg}"
            )

            retry_choice = messagebox.askretrycancel(
                "Cache Save Error",
                f"Failed to save SMTP cache:\n{e}\n\n"
                "Click Retry to try again, or Cancel for TXT fallback."
            )

            if not retry_choice:
                txt_choice = messagebox.askyesno(
                    "Save Cache as TXT",
                    f"Save new entries as TXT to:\n{txt_fallback_dir}\n\n"
                    "Each line: EntryName<TAB>SMTPAddress"
                )
                if txt_choice:
                    try:
                        os.makedirs(txt_fallback_dir, exist_ok=True)
                        with open(txt_fallback_path, "a", encoding="utf-8") as f:
                            for entry_name, smtp_address in self.new_smtp_entries.items():
                                f.write(f"{entry_name}\t{smtp_address}\n")
                        self.invalid_logger.info(
                            f"SMTPCacheSaveFallback|save_smtp_cache|"
                            f"Saved {len(self.new_smtp_entries)} entries to TXT: {txt_fallback_path}"
                        )
                        print(
                            f"Saved {len(self.new_smtp_entries)} SMTP entries "
                            f"to TXT fallback: {txt_fallback_path}"
                        )
                        self.new_smtp_entries.clear()
                    except Exception as txt_e:
                        self.invalid_logger.error(
                            f"CacheSaveFallbackError||save_smtp_cache|{txt_e}"
                        )
                break   # exit loop regardless of TXT outcome

            time.sleep(1)   # brief pause before manual retry
